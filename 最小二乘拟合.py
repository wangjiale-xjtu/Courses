import numpy as np
import openpyxl



def SolvingLinearEquations(AugmentedMatrix):
    col = AugmentedMatrix.shape[1] #增广矩阵列数
    #消元
    for i in range(col- 2):
        current_column  = AugmentedMatrix[i:,i]
        max_index = np.argmax(current_column) + i #寻找最大元
        if(AugmentedMatrix[max_index,i] == 0):
            print("无唯一解")
            return
        AugmentedMatrix[[i,max_index],:] = AugmentedMatrix[[max_index,i],:] #交换
        l = AugmentedMatrix[i+1:,i] / AugmentedMatrix[i,i] #计算系数
        m =  np.tile(AugmentedMatrix[i,:],(l.shape[0],1)) * np.tile(l,(col,1)).T #计算消元时减去的矩阵
        AugmentedMatrix[i+1:,:] = AugmentedMatrix[i+1:,:] - m #消元
    if(AugmentedMatrix[col - 2,col - 2] == 0):
            print("无唯一解")
            return
   #代入
    x = np.zeros(col-1)
    for i in range(col-2,-1,-1):
        x[i] = (AugmentedMatrix[i,-1] - np.dot(AugmentedMatrix[i,:-1] , x.T)) / AugmentedMatrix[i,i]
    return x


def regularformulation(X, Y, Degree):
    A=np.zeros([Degree+1,Degree+2])
    X=np.array([X])
    Y = np.array([Y])
    for i in range(Degree+1):
        A[i,Degree+1]=sum([(x[0]**i)*x[1] for x in np.hstack((X.T,Y.T))])

        for j in range(Degree+1):
            A[i,j]=sum([x**(i+j) for x in np.ravel(X)])


    return A




if __name__ == '__main__':      # 模块被直接运行时，以下代码块将被运行，当模块是被导入时，代码块不被运行。
    ##下面这一段可以实现从excel中读取数据
    '''
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']
    N_i = ws.max_column
    X_i=[]
    Y_i=[]
    for i in range(N_i):
        cell1 = ws.cell(row=1, column=i)  
        cell2 = ws.cell(row=2, column=i) 
        X_i.append(int(cell2.value))
        Y_i.append(int(cell1.value))
    '''
    Degree = 9  # 多项式次数可变
    X_i = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]  # 拟合的数据可变
    Y_i = [5.1234, 5.3057, 5.5678, 5.9375, 6.4370, 7.0978, 7.9493, 9.0253, 10.3627]
    a = np.array([[1, 2, 1, 0], [2, 2, 3, 3], [-1, -3, 0, 2]])
    A=regularformulation(X_i,Y_i,Degree)
    print(A)
    c=SolvingLinearEquations(A)
    c=np.array([c])
    print(SolvingLinearEquations(A))
    Y_pre = []
    order=list(range(Degree + 1))
    # order.reverse()
    order=np.array([order])
    for i in range(len(X_i)):
        Y_pre.append(sum([x[0]*(X_i[i]**x[1]) for x in np.hstack((c.T,order.T))]))
    print(Y_pre)
    error=0
    for i in range(len(X_i)):
        error=error+abs((Y_pre[i]-Y_i[i])/Y_i[i])/len(X_i)
    print('平均相对误差为'+str(error))