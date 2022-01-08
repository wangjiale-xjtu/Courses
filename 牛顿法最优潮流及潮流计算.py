import numpy as np
from numpy.lib.arraypad import _get_stats
import openpyxl
import cvxpy as cvx
import math

def NRPolar(U,Angle,Y,PQNode,SlackNode,P_Real,Q_Real,Tol,**Option):
    print()
    P_iter=0
    Q_iter=0
    PQNode=[x-1 for x in PQNode]
    SlackNode=SlackNode-1
    NumNode=Y.shape[0]
    NumPQ=len(PQNode)
    G=Y.real
    B=Y.imag 
    P=np.zeros((NumNode,1))  
    Q=np.zeros((NumNode,1)) 
    DeltaP=np.zeros((NumNode-1,1))
    DeltaQ=np.zeros((NumPQ,1))
    print(DeltaP.shape)
    print(DeltaQ.shape)
    print(P_Real.shape)
    print(P.shape)
    for i in range(NumNode): #求解功率不平衡量
        P[i] = U[i]*np.sum(U*(G[i,:]*np.cos(Angle[i]-Angle) +  B[i,:]*np.sin(Angle[i]-Angle)))
        Q[i] = U[i]*np.sum(U*(G[i,:]* np.sin(Angle[i]-Angle) -  B[i,:]*np.cos(Angle[i]-Angle)))
        if i!=SlackNode:
            DeltaP[P_iter]=P_Real[i]-P[i]
            P_iter+=1
            if i in PQNode:
                DeltaQ[Q_iter]=Q_Real[i]-Q[i]
                Q_iter+=1
        
    DeltaPQ=np.vstack([DeltaP,DeltaQ])
    #print('功率不平衡量为\n',DeltaPQ)
    MaxError=np.max(np.abs(DeltaPQ))
    print(MaxError)
    if MaxError<Tol:
        return (U,Angle,MaxError)
    HN_iter = -1   # 初始化雅可比矩阵
    H = np.zeros([NumNode-1,NumNode-1])
    N = np.zeros([NumNode-1,NumPQ])
    # H and N
    for i in range(NumNode):
        if i!=SlackNode:  # PQ或PV节点
            H_iter_y = -1 
            N_iter_y = -1
            HN_iter = HN_iter+1  # 记录H和N的行数
            for j in range(NumNode):
                if j != SlackNode:
                    H_iter_y = H_iter_y+1  # 记录H列数
                    if i != j:   # 非平衡节点计算H矩阵
                        Angleij = Angle[i]-Angle[j]
                        H[HN_iter,H_iter_y] = -U[i]*U[j]*(G[i,j]*np.sin(Angleij)-B[i,j]*np.cos(Angleij))
                    else:
                        H[HN_iter,H_iter_y] = Q[i]+U[i]**2*B[i,i]
                    if j in PQNode:
                        N_iter_y = N_iter_y+1  # 记录N的列数
                        if i != j:
                            Angleij = Angle[i]-Angle[j]
                            N[HN_iter,N_iter_y] = -U[i]*U[j]*(G[i,j]*np.cos(Angleij)+B[i,j]*np.sin(Angleij))
                        else:
                            N[HN_iter,N_iter_y] = -P[i]-G[i,i]*U[i]**2
    ## J and L
    JL_iter = -1   # 初始化雅可比矩阵
    J = np.zeros([NumPQ,NumNode-1])
    L = np.zeros([NumPQ,NumPQ])
    for i in range(NumNode):
        if i in PQNode:    # PQ节点
            JL_iter = JL_iter+1 # J和L的行数
            J_iter_y = -1
            L_iter_y = -1
            for j in range(NumNode):
                if j!=SlackNode:  # 非平衡节点
                    J_iter_y = J_iter_y+1
                    if i!=j:
                        Angleij = Angle[i]-Angle[j]
                        J[JL_iter,J_iter_y] = U[i]*U[j]*(G[i,j]*np.cos(Angleij)+B[i,j]*np.sin(Angleij))
                    else:
                        J[JL_iter,J_iter_y] = -P[i]+G[i,i]*U[i]**2
                    if j in PQNode:  # PQ节点
                        L_iter_y = L_iter_y+1
                        if i!=j:
                            Angleij = Angle[i]-Angle[j]
                            L[JL_iter,L_iter_y] = -U[i]*U[j]*(G[i,j]*np.sin(Angleij)-B[i,j]*np.cos(Angleij))
                        else:
                            L[JL_iter,L_iter_y] = -Q[i]+B[i,i]*U[i]**2
    # 修正
    Jaccobi = np.vstack([np.hstack([H,N]),np.hstack([J,L])])
    Option['string'] = 'jacobi矩阵为：\n'
    
    Delta = np.linalg.solve(Jaccobi,DeltaPQ)
    Option['string'] = '方程组求解结果：\n'
    
    DeltaAngle = Delta[0:NumNode-1]  # 注意下标
    DeltaU_U = Delta[NumNode-1:]
    DA_iter = -1
    U_U_iter = -1
    for i in range(NumNode):
        if i!=SlackNode:
            DA_iter = DA_iter+1
            Angle[i] = Angle[i]-DeltaAngle[DA_iter]
            if i in PQNode:
                U_U_iter = U_U_iter+1
                U[i] = U[i]-U[i]*DeltaU_U[U_U_iter]
    Option['string'] = '更新之后的电压幅值为：\n'
    
    Option['string'] = '相角为：\n'
    
    return(U,Angle,MaxError)




def OutData(U,Angle,P,Q,location,V_thetaNode):
    wb=openpyxl.load_workbook(location) 
    ws=wb['Sheet1']#相应工作表
    n=ws.max_row-1
    for t in range(n):
        cell1=ws.cell(row=t+2,column=9) #U
        cell2=ws.cell(row=t+2,column=10) #Angle
        cell5=ws.cell(row=t+2,column=11) #节点类型
        cell3=ws.cell(row=t+2,column=12) #P
        cell4=ws.cell(row=t+2,column=13) #Q
        cell1.value=U[t,0]
        cell2.value=Angle[t,0]
        cell3.value=P[t,0]
        cell4.value=Q[t,0] 
        if t==V_thetaNode-1:
            cell5.value=3
        else:    
            cell5.value=1      
    wb.save(location)

#得到节点导纳矩阵
def getY(filename,Sbase,accuracy):
    wb=openpyxl.load_workbook(filename+'\\line.xlsx')
    wb2=openpyxl.load_workbook(filename+'\\bus.xlsx')
    ws=wb['Sheet1']
    ws2=wb2['Sheet1']
    #busrow=ws2.max_row-1
    branchnum=ws.max_row-1
    nodelist=[]
    for i in range(branchnum):
        cell1=ws.cell(row=i+2,column=2)  #from
        cell2=ws.cell(row=i+2,column=3)  #to
        nodelist.append(int(cell1.value))
        nodelist.append(int(cell2.value))
    busnum=max(nodelist)
    Y=np.zeros((busnum,busnum))+1j*np.zeros((busnum,busnum))
    for i in range(busnum):
        cell1=float(ws2.cell(row=i+2,column=7).value)  #Gs
        cell2=float(ws2.cell(row=i+2,column=8).value)  #Bs
        cell3=int(ws2.cell(row=i+2,column=1).value)    #bus
        Y[cell3-1,cell3-1]=Y[cell3-1,cell3-1]+cell1/Sbase+1j*cell2/Sbase
    
    for i in range(branchnum):
        fbus=int(ws.cell(row=i+2,column=2).value)  #from
        tbus=int(ws.cell(row=i+2,column=3).value)   #to
        X=float(ws.cell(row=i+2,column=4).value)  #X
        R=float(ws.cell(row=i+2,column=6).value)   #R
        B=float(ws.cell(row=i+2,column=8).value)   #pi型等值对地电容
        Ratio=float(ws.cell(row=i+2,column=9).value)#变压器变比，若非变压器支路则为0
        #nodeteam.append((int(cell1.value),int(cell2.value)))
        if Ratio==0: #非变压器支路
            Y[fbus-1,tbus-1]=-1/(R+1j*X)
            Y[tbus-1,fbus-1]=-1/(R+1j*X)
            Y[fbus-1,fbus-1]=Y[fbus-1,fbus-1]+1/(R+1j*X)+1j*B
            Y[tbus-1,tbus-1]=Y[tbus-1,tbus-1]+1/(R+1j*X)+1j*B
        else:  #变压器支路
            Y0=1/(R+1j*X)
            Y[fbus-1,tbus-1]=-1/Ratio*Y0
            Y[tbus-1,fbus-1]=-1/Ratio*Y0
            Y[fbus-1,fbus-1]=Y[fbus-1,fbus-1]+1/Ratio*Y0+(1-Ratio)/Ratio/Ratio*Y0
            Y[tbus-1,tbus-1]=Y[tbus-1,tbus-1]+(Ratio-1)/Ratio*Y0+Y0/Ratio 
        Y=np.round(Y,accuracy)     
    return Y

def getNetdata(filename):
    PQNode=[]
    wb2=openpyxl.load_workbook(filename+'\\bus.xlsx')
    ws2=wb2['Sheet1']
    busrow=ws2.max_row-1
    P_Injection=np.zeros(busrow)
    Q_Injection=np.zeros(busrow)
    for i in range(busrow):
        kind=int(ws2.cell(row=i+2,column=11).value)
        if kind==1:
            PQNode.append(int(ws2.cell(row=i+2,column=1).value))
        elif kind==3:
            SlackNode=int(ws2.cell(row=i+2,column=1).value)
        P_Injection[i]=float(ws2.cell(row=i+2,column=12).value)
        Q_Injection[i]=float(ws2.cell(row=i+2,column=13).value)
    return PQNode,SlackNode,P_Injection,Q_Injection

def getU_Angle0(filename):
    wb2=openpyxl.load_workbook(filename+'\\bus.xlsx')
    ws2=wb2['Sheet1']
    busrow=ws2.max_row-1
    U=np.zeros(busrow)
    Angle=np.zeros(busrow)
    for i in range(busrow):
        U[i,]=float(ws2.cell(row=i+2,column=9).value)  #U
        Angle[i,]=float(ws2.cell(row=i+2,column=10).value)  #angle
    return U,Angle


def getUsimilar(W):
    #奇异值分解
    n=W.shape[0]
    u, SingularValue, vh = np.linalg.svd(W)
    #print('分解得到矩阵的形状：\n',u.shape,s.shape,vh.shape)
 
    #print('奇异值：\n',s)
 
    smat=np.zeros((n,n))
 
    smat[:n,:n]= np.diag(SingularValue)
 
    #print('奇异矩阵：\n',smat)
    #print('u：\n',u)
    #print('vh：\n',vh)
    #print(u/vh.T)
    #print(W.value/(u@smat@vh))
    smatsimilar=np.zeros((n,n))
    smatsimilar[0,0]=smat[0,0]
    Wsimilar=np.zeros((n,n))
    a=np.zeros((n,1),dtype=complex)
    b=np.zeros((1,n),dtype=complex)
    c=np.zeros((1,1),dtype=complex)
    for i in range(n):
        a[i,0]=u[i,0]
        b[0,i]=vh[0,i]
    c[0,0]=smat[0,0]
    WS=a@c@b
    #print(a@c@b)
    #print(np.linalg.matrix_rank(WS))
    Usimilar=-(a+b.conjugate().T)/2*(smat[0,0])**0.5  #我们认为的近似的电压
    return Usimilar,SingularValue

def getU_Angle(Usimilar):
    U=np.zeros((np.size(Usimilar),1))
    Angle=np.zeros((np.size(Usimilar),1))
    for i in range(np.size(Usimilar)):
        U[i]=abs(Usimilar[i])
        Angle[i]=math.atan(np.imag(Usimilar[i])/np.real(Usimilar[i]))
    return U,Angle




def get_otherData(menu0,Y):
    menu1=menu0+'\\generator.xlsx'
    menu2=menu0+'\\line.xlsx'
    menu3=menu0+'\\bus.xlsx'
    wb1=openpyxl.load_workbook(menu1)
    wb2=openpyxl.load_workbook(menu2)
    wb3=openpyxl.load_workbook(menu3)
    ws1=wb1['Sheet1']
    ws2=wb2['Sheet1']
    ws3=wb3['Sheet1']
    n1=ws1.max_row-1  #发电机数
    n2=ws2.max_row-1 #支路数
    n3=ws3.max_row-1  #节点数
    nodelist=[]
    #找出节点的最大序号，即总数
    for i in range(n2):
        cell1=ws2.cell(row=i+2,column=2)
        cell2=ws2.cell(row=i+2,column=3)
        nodelist.append(cell1.value)
        nodelist.append(cell2.value)
    nodenum=max(nodelist)   #按理说和n3是一样的
    #线路节点对
    nodeteam=[]
    #线路、节点关联矩阵
    K=np.zeros((nodenum,n2))
    #线路和线路起点\终点的关联矩阵
    Ki=np.zeros((nodenum,n2))
    Kj=np.zeros((nodenum,n2))
    #线路总容量
    Sijmax=np.zeros((n2,1))
    ##b矩阵和g矩阵
    g=np.zeros((n2,1))
    b=np.zeros((n2,1))
    #线路和起点的两种关联矩阵
    Kij=np.zeros((nodenum,n2))
    Kji=np.zeros((nodenum,n2))
    for i in range(n2):
        cell1=ws2.cell(row=i+2,column=2)  #from
        cell2=ws2.cell(row=i+2,column=3)  #to
        cell3=ws2.cell(row=i+2,column=4)  #X
        cell4=ws2.cell(row=i+2,column=6)  #R
        nodeteam.append((int(cell1.value),int(cell2.value)))
        X0=float(cell3.value)   #标幺值
        R0=float(cell4.value)   #标幺值

        Kij[int(cell1.value)-1,i]=1
        Kji[int(cell2.value)-1,i]=1
    
        K[int(cell1.value)-1,i]=1
        Ki[int(cell1.value)-1,i]=1
        K[int(cell2.value)-1,i]=-1
        Kj[int(cell2.value)-1,i]=1
        Sijmax[i]=float(ws2.cell(row=i+2,column=7).value)/100
    
        g[i,]=R0/(R0*R0+X0*X0) 
        b[i,]=X0/(R0*R0+X0*X0)
    ##导纳矩阵生成完毕
    print('g:\n',g)
    print('b:\n',b)
    #print('线路容量：\n',Sijmax)
    print('nodeteam:\n',nodeteam)
    #节点-机组关联矩阵
    node_gene=np.zeros((nodenum,n1))
    #自导纳矩阵
    Yii=np.zeros((n3,1),dtype=complex)
    #节点电压上下限
    Vmax=np.zeros((nodenum,1))
    Vmin=np.zeros((nodenum,1))
    LP=np.zeros((nodenum,1))   #有功负荷
    LQ=np.zeros((nodenum,1))   #无功负荷
    for i in range(n3):
        cell1=ws3.cell(row=i+2,column=2)  #有功负荷
        cell2=ws3.cell(row=i+2,column=3)  #无功负荷
        cell3=ws3.cell(row=i+2,column=4)  #电压上限
        cell4=ws3.cell(row=i+2,column=5)  #电压下限
        Vmax[i,]=float(cell3.value)  #标幺值
        Vmin[i,]=float(cell4.value)  #标幺值
        LP[i,]=float(cell1.value)/Sbase
        LQ[i,]=float(cell2.value)/Sbase
        Yii[i]=np.sum(Y[i,:])
    Gs=np.zeros((n3,1))
    Bs=np.zeros((n3,1))
    Gs=np.real(Yii)
    Bs=np.imag(Yii)
    print('并联导纳:\n',Yii)
    print('并联电导:\n',Gs)
    print('并联电纳:\n',Bs)
    #机组发电成本
    ag=np.zeros((n1,1))
    a1=[]
    bg=np.zeros((n1,1))
    cg=np.zeros((n1,n1))
    cgg=np.zeros((n1,1))
    #机组上下限
    PGmin=np.zeros((n1,1))
    PGmax=np.zeros((n1,1))
    QGmin=np.zeros((n1,1))
    QGmax=np.zeros((n1,1))
    for i in range(n1):
        cell1=ws1.cell(row=i+2,column=2)
        cell2=ws1.cell(row=i+2,column=9)
        cell3=ws1.cell(row=i+2,column=10)
        cell4=ws1.cell(row=i+2,column=11)
        cell5=ws1.cell(row=i+2,column=3)
        cell6=ws1.cell(row=i+2,column=4)
        cell7=ws1.cell(row=i+2,column=13)
        cell8=ws1.cell(row=i+2,column=12)
        node_gene[int(cell1.value)-1,i]=1
        a1.append(float(cell2.value))
        ag[i]=float(cell2.value)
        bg[i]=float(cell3.value)
        cgg[i]=float(cell4.value)
        cg[i,i]=float(cell4.value)
        PGmin[i,]=float(cell5.value)/Sbase
        PGmax[i,]=float(cell6.value)/Sbase
        QGmin[i,]=float(cell7.value)/Sbase
        QGmax[i,]=float(cell8.value)/Sbase
    diaga=np.diag(a1)
    print('节点机组关联矩阵\n',node_gene)
    print('bg:\n',bg)
    print('cg:\n',cg)
    print('发电机最小有功出力PGmin:\n',PGmin)
    print('发电机最大有功出力PGmax:\n',PGmax)
    print('发电机最小无功出力QGmin:\n',QGmin)
    print('发电机最大无功出力QGmax:\n',QGmax)
    print('有功负荷LP:\n',LP)
    print('无功负荷LQ:\n',LQ)
    y=np.zeros((n2,1),dtype=complex)
    for i in range(n2):
        y[i]=g[i]+1j*b[i]
    return Yii,nodenum,n3,n2,n1,cg,bg,cgg,ag,nodeteam,Sijmax,Vmax,Vmin,Kij,Kji,Bs,Gs,node_gene,LP,LQ,PGmax,PGmin,QGmax,QGmin

def Optimization(delta): 
    One=np.array([[1,0],[0,1]])
    One1=np.ones((nodenum,1))
    Two=np.array([[2,0,0,0],[0,2,0,0],[0,0,1,-1]])
    Two1=np.array([0,0,1,1])
    test1=np.array([[1,1,0,0],[0,0,1,-1]])
    test2=np.array([1,1,1,1])
    Pij=cvx.Variable((n2,1))
    Pji=cvx.Variable((n2,1))
    Qij=cvx.Variable((n2,1))
    Qji=cvx.Variable((n2,1))
    #Sij=cvx.Variable((n2,1),complex=True)
    #Sji=cvx.Variable((n2,1),complex=True)
    Pg=cvx.Variable((n1,1))
    Qg=cvx.Variable((n1,1))
    W=cvx.Variable((nodenum,nodenum),hermitian=True)
    p=cgg.T@(Pg*100)**2+bg.T@(Pg*100)+cvx.sum(ag)
    #obj = cvx.Minimize(p-0*cvx.sum([cvx.real(W[i-1,j-1]) for (i,j) in nodeteam]))
    #obj = cvx.Minimize(p-0.00001*cvx.sum([cvx.real(W[i-1,j-1]) for (i,j) in nodeteam]))
    #obj = cvx.Minimize(p-0.001*cvx.sum([cvx.real(W[i-1,j-1]) for (i,j) in nodeteam]))
    obj = cvx.Minimize(cgg.T@(Pg*100)**2+bg.T@(Pg*100)+cvx.sum(ag)+delta*cvx.sum(Qg))
    con=[]
    for i in range(n2):
        #con+=[Sij[i]==y[i]*(W[nodeteam[i][0]-1,nodeteam[i][0]-1]-W[nodeteam[i][0]-1,nodeteam[i][1]-1])]
        #con+=[Sji[i]==y[i]*(W[nodeteam[i][1]-1,nodeteam[i][1]-1]-W[nodeteam[i][1]-1,nodeteam[i][0]-1])]
        con+=[Pij[i]+1j*Qij[i]==(-Y[nodeteam[i][0]-1,nodeteam[i][1]-1].conjugate())*(W[nodeteam[i][0]-1][nodeteam[i][0]-1]-W[nodeteam[i][0]-1][nodeteam[i][1]-1])]
        con+=[Pji[i]+1j*Qji[i]==(-Y[nodeteam[i][1]-1,nodeteam[i][0]-1].conjugate())*(W[nodeteam[i][1]-1][nodeteam[i][1]-1]-W[nodeteam[i][1]-1][nodeteam[i][0]-1])]
        #con+=[cvx.SOC(Sijmax[i],One@cvx.vstack([Pij[i],Qij[i]]))]
        #con+=[cvx.SOC(Sijmax[i],One@cvx.vstack([Pji[i],Qji[i]]))]
    for i in range(nodenum):
        con+=[cvx.real(W[i,i])<=Vmax[i]**2]
        con+=[cvx.real(W[i,i])>=Vmin[i]**2]
    con+=[W >> 0] 
    #for j in range(5):
        #for k in range(5):
            #zjbl=cvx.vstack((cvx.real(W[i-1][j-1]),cvx.imag(W[i-1][j-1]),cvx.real(W[i-1][i-1]),cvx.real(W[j-1][j-1])))
            #con+=[cvx.SOC(test1 @ zjbl,test2@zjbl)]
    #con+=[Kij@Sij+Kji@Sji==node_gene@(Pg+1j*Qg)-(LP+1j*LQ)]
    #for i in range(n3):
        #con+=[1j*(Kij@Qij+Kji@Qji)[i]+(Kij@Pij)[i]+(Kji@Pji)[i]+(cvx.multiply(Yii,cvx.vstack([W[i,i] for i in range(nodenum)])))[i]==(node_gene@Pg)[i]-LP[i]+1j*(node_gene@Qg-LQ)[i]]  
    #con+=[Kij@Qij+Kji@Qji+cvx.multiply(Bs,cvx.vstack([W[i,i] for i in range(nodenum)]))==node_gene@Qg-LQ]
    #con+=[Kij@Pij+Kji@Pji+cvx.multiply(Gs,cvx.vstack([W[i,i] for i in range(nodenum)]))==node_gene@Pg-LP]
    for i in range(n3): 
        #con+=[(Kij@Qij)[i]+(Kji@Qji)[i]==(node_gene@Qg)[i]-LQ[i]]
        #con+=[(Kij@Pij)[i]+(Kji@Pji)[i]==(node_gene@Pg)[i]-LP[i]]
        con+=[(Kij@Qij)[i]+(Kji@Qji)[i]-Bs[i]*W[i,i]==(node_gene@Qg)[i]-LQ[i]]
        con+=[(Kij@Pij)[i]+(Kji@Pji)[i]+Gs[i]*W[i,i]==(node_gene@Pg)[i]-LP[i]]  
    con+=[ 
        Pg<=PGmax,
        Qg<=QGmax,
        Pg>=PGmin,
        Qg>=QGmin       
        ]

    prob=cvx.Problem(obj,con)
    mydict = {}
    #mydict = {"MSK_DPAR_BASIS_REL_TOL_S":1e-0,"MSK_DPAR_BASIS_TOL_S":1e-0,"MSK_DPAR_BASIS_TOL_X":1e-0,"MSK_IPAR_BI_IGNORE_NUM_ERROR":1}  
    prob.solve(solver='MOSEK')
    return Pij.value,Pji.value,Qij.value,Qji.value,W.value,Pg.value,Qg.value,obj.value,prob.status,p.value


def PirntFun():
    print("节点导纳矩阵为：\n",Y)
    print('问题求解状态:\n',status)
    print('发电机有功出力Pg:\n',Pg)
    print('*****************************************')
    print('线路正向有功Pij:\n',Pij)
    print('*****************************************')
    print('线路反向有功Pji:\n',Pji)
    print('*****************************************')
    print('线路正向无功Qij:\n',Qij)
    print('*****************************************')
    print('线路反向无功Qji:\n',Qji)
    print('*****************************************')    
    print('节点总注入无功功率:\n',node_gene@Pg-LP)
    print('*****************************************')
    print('节点总注入无功功率:\n',node_gene@Qg-LQ)
    print('*****************************************')
    print(W)
    print('*****************************************')
    print('''W的秩为''',np.linalg.matrix_rank(W))
    print('''目标函数值为''',obj)
    print('''p值为''',p)
    #print('''特征值向量为''',e_vals)
    #print('''特征向量矩阵为''',e_vecs)
    #print('奇异值近似电压幅值为:',U)
    #print('奇异值近似电压相角为:',Angle)
    print('30节点奇异值为：\n',SingularValue)
    print('目标函数值为：\n',obj)
    #print("特征值：", eigenvalue)
    #print("特征向量：", featurevector)
    print('并联导纳:\n',Yii)
    print('并联电导:\n',Gs)
    print('并联电纳:\n',Bs)
    print('近似电压Usimilar:\n',Usimilar)

def PowerFlowCal(Flag_Out):
    # 初始化
    U,Angle = getU_Angle0(menu0)
    PQNode,SlackNode,P_Real,Q_Real = getNetdata(menu0)

    # 开始计算
    Iter = 0
    MaxIter = 10
    Tol = 1e-10

    while True:
        Iter = Iter+1
        U,Angle,MaxError = NRPolar(U,Angle,Y,PQNode,SlackNode,P_Real,Q_Real,Tol,width=9)
        if Iter>MaxIter or MaxError<Tol:
            break
    # 结束循环，输出结果
    if Flag_Out==1:
        if MaxError<Tol:
            print('迭代完成，迭代次数为：',Iter)
            print('最大误差为：',MaxError)
            print('电压幅值为：',U)
            print('电压相角为：',Angle)    
        else:
            print('结果不收敛!')
    wb=openpyxl.load_workbook(menu0+'\\bus.xlsx') 
    ws=wb['Sheet1']#相应工作表
    n=ws.max_row-1
    for t in range(n):
        cell1=ws.cell(row=t+2,column=15) #U
        cell2=ws.cell(row=t+2,column=16) #Angle
        cell1.value=U[t,]
        cell2.value=Angle[t,]     
        
    Pinj=0
    Qinj=0
    for i in range(U.size):
        Pinj+=U[0]*U[i]*(Y[0,i].real*np.cos(Angle[0]-Angle[i])+Y[0,i].imag*np.sin(Angle[0]-Angle[i]))
        Qinj+=U[0]*U[i]*(Y[0,i].real*np.sin(Angle[0]-Angle[i])-Y[0,i].imag*np.cos(Angle[0]-Angle[i]))
    cell1=ws.cell(row=2,column=32)
    cell2=ws.cell(row=2,column=33)
    cell1.value=Pinj
    cell2.value=Qinj
    wb.save(menu0+'\\bus.xlsx')
################################################################################################################################




#基准值
Sbase=100
Vbase=115
Zbase=Vbase*Vbase/Sbase
Ybase=1/Zbase
accuracy=10    #节点导纳矩阵的保留位数
Flag_Out=0
delta=0
menu0='C:\\Users\\2020\\Desktop\\case30ww'
menu1=menu0+'\\generator.xlsx'
menu2=menu0+'\\line.xlsx'
menu3=menu0+'\\bus.xlsx'
Y=getY(menu0,Sbase,accuracy)  ##节点导纳矩阵
Yii,nodenum,n3,n2,n1,cg,bg,cgg,ag,nodeteam,Sijmax,Vmax,Vmin,Kij,Kji,Bs,Gs,node_gene,LP,LQ,PGmax,PGmin,QGmax,QGmin=get_otherData(menu0,Y)
Pij,Pji,Qij,Qji,W,Pg,Qg,obj,status,p0=Optimization(delta)
print("---------delta=",delta,"---------")
print("成本函数差额为",0)
Usimilar,SingularValue=getUsimilar(W)
print("奇异值为：")
print(SingularValue)
#p=np.zeros(30)
SV=[]
sta=[]
print(obj)
mubiaof=[]
chengbenf=[]
delta=-5
'''
for i in range(30):
    delta=delta+5
    try:
        Pij,Pji,Qij,Qji,W,Pg,Qg,obj,status,p[i]=Optimization(delta)
        Usimilar,SingularValue=getUsimilar(W)
    except:
        SingularValue=np.zeros(2)
        SingularValue[1]=1
    SV.append(SingularValue[0]/SingularValue[1])
    sta.append(status)
    mubiaof.append(obj)

    print("第"+str(i)+"次完成")
    
delta=-5
for i in range(30):
    delta+=5
    print("---------delta=",delta,"---------")
    print("求解状态：",sta[i])
    print("成本函数为",p[i])
    print("目标函数为",mubiaof[i])
    print("奇异值比值为：")
    print(math.log(SV[i],10))

'''
'''
Pinjection=node_gene@Pg-LP
Qinjection=node_gene@Qg-LQ
e_vals,e_vecs = np.linalg.eig(W)
Usimilar,SingularValue=getUsimilar(W)
U,Angle=getU_Angle(Usimilar)
OutData(U,Angle,Pinjection,Qinjection,menu3,1)
eigenvalue, featurevector = np.linalg.eig(W)
PowerFlowCal(Flag_Out)
'''
#PirntFun()
#print("原模型奇异值之比为",SingularValue[0]/SingularValue[1])
#print(SingularValue)
T_tF = 200 # tao
u = 2
T_max = 1e5
#T_max = 1e5
tF = 0
delta_e = 1e-6
#delta_e = 1e-6
nt=nodeteam
grad=np.zeros((n2,4))
M=W
'''
'''
At=[]
for i in range(nodenum):
    for j in range(i+1,nodenum):
        a=(i,j)
        At.append(a)
while tF<=100:
    Wr_tF=M.real
    Wi_tF=M.imag  
    for i in range(nodenum):
        grad[i,:]=[-2*Wr_tF[At[i][0]-1,At[i][1]-1],-2*Wi_tF[At[i][0]-1,At[i][1]-1],Wr_tF[At[i][1]-1,At[i][1]-1],Wi_tF[At[i][1]-1,At[i][1]-1]]
    One=np.array([[1,0],[0,1]])
    One1=np.ones((nodenum,1))
    Two=np.array([[2,0,0,0],[0,2,0,0],[0,0,1,-1]])
    Two1=np.array([0,0,1,1])
    test1=np.array([[1,1,0,0],[0,0,1,-1]])
    test2=np.array([1,1,1,1])
    Pij=cvx.Variable((n2,1))
    Pji=cvx.Variable((n2,1))
    Qij=cvx.Variable((n2,1))
    Qji=cvx.Variable((n2,1))
    #Sij=cvx.Variable((n2,1),complex=True)
    #Sji=cvx.Variable((n2,1),complex=True)
    Pg=cvx.Variable((n1,1))
    Qg=cvx.Variable((n1,1))
    W=cvx.Variable((nodenum,nodenum),hermitian=True)
    sigma=cvx.Variable((nodenum,1))
    p=cgg.T@(Pg*100)**2+bg.T@(Pg*100)+cvx.sum(ag)
    #obj = cvx.Minimize(p-0*cvx.sum([cvx.real(W[i-1,j-1]) for (i,j) in nodeteam]))
    #obj = cvx.Minimize(p-0.00001*cvx.sum([cvx.real(W[i-1,j-1]) for (i,j) in nodeteam]))
    #obj = cvx.Minimize(p-0.001*cvx.sum([cvx.real(W[i-1,j-1]) for (i,j) in nodeteam]))
    obj = cvx.Minimize(cgg.T@(Pg*100)**2+bg.T@(Pg*100)+T_tF*cvx.sum(sigma))
    con=[]
    for i in range(n2):
        #con+=[Sij[i]==y[i]*(W[nodeteam[i][0]-1,nodeteam[i][0]-1]-W[nodeteam[i][0]-1,nodeteam[i][1]-1])]
        #con+=[Sji[i]==y[i]*(W[nodeteam[i][1]-1,nodeteam[i][1]-1]-W[nodeteam[i][1]-1,nodeteam[i][0]-1])]
        con+=[Pij[i]+1j*Qij[i]==(-Y[nodeteam[i][0]-1,nodeteam[i][1]-1].conjugate())*(W[nodeteam[i][0]-1][nodeteam[i][0]-1]-W[nodeteam[i][0]-1][nodeteam[i][1]-1])]
        con+=[Pji[i]+1j*Qji[i]==(-Y[nodeteam[i][1]-1,nodeteam[i][0]-1].conjugate())*(W[nodeteam[i][1]-1][nodeteam[i][1]-1]-W[nodeteam[i][1]-1][nodeteam[i][0]-1])]
        #con+=[cvx.SOC(Sijmax[i],One@cvx.vstack([Pij[i],Qij[i]]))]
        #con+=[cvx.SOC(Sijmax[i],One@cvx.vstack([Pji[i],Qji[i]]))]
    for i in range(nodenum):
        con+=[cvx.real(W[i,i])<=Vmax[i]**2]
        con+=[cvx.real(W[i,i])>=Vmin[i]**2]
    con+=[W >> 0] 
    #for j in range(5):
        #for k in range(5):
            #zjbl=cvx.vstack((cvx.real(W[i-1][j-1]),cvx.imag(W[i-1][j-1]),cvx.real(W[i-1][i-1]),cvx.real(W[j-1][j-1])))
            #con+=[cvx.SOC(test1 @ zjbl,test2@zjbl)]
    #con+=[Kij@Sij+Kji@Sji==node_gene@(Pg+1j*Qg)-(LP+1j*LQ)]
    #for i in range(n3):
        #con+=[1j*(Kij@Qij+Kji@Qji)[i]+(Kij@Pij)[i]+(Kji@Pji)[i]+(cvx.multiply(Yii,cvx.vstack([W[i,i] for i in range(nodenum)])))[i]==(node_gene@Pg)[i]-LP[i]+1j*(node_gene@Qg-LQ)[i]]  
    #con+=[Kij@Qij+Kji@Qji+cvx.multiply(Bs,cvx.vstack([W[i,i] for i in range(nodenum)]))==node_gene@Qg-LQ]
    #con+=[Kij@Pij+Kji@Pji+cvx.multiply(Gs,cvx.vstack([W[i,i] for i in range(nodenum)]))==node_gene@Pg-LP]
    for i in range(n3): 
        con+=[(Kij@Qij)[i]+(Kji@Qji)[i]-Bs[i]*W[i,i]==(node_gene@Qg)[i]-LQ[i]]
        con+=[(Kij@Pij)[i]+(Kji@Pji)[i]+Gs[i]*W[i,i]==(node_gene@Pg)[i]-LP[i]]  
    con+=[ 
        Pg<=PGmax,
        Qg<=QGmax,
        Pg>=PGmin,
        Qg>=QGmin       
        ]
    ##CCP额外约束
    for i in range(nodenum):
        con+=[sigma[i,0]>=0]
        con+=[Wr_tF[At[i][0]-1,At[i][0]-1]*Wr_tF[At[i][1]-1,At[i][1]-1]-Wr_tF[At[i][0]-1,At[i][1]-1]**2-Wi_tF[At[i][0]-1,At[i][1]-1]**2+grad[i,:]@cvx.vstack([cvx.real(W[At[i][0]-1,At[i][1]-1])-Wr_tF[At[i][0]-1,At[i][1]-1],cvx.imag(W[At[i][0]-1,At[i][1]-1])-Wi_tF[At[i][0]-1,At[i][1]-1],cvx.real(W[At[i][0]-1,At[i][0]-1])-Wr_tF[At[i][0]-1,At[i][0]-1],cvx.real(W[At[i][1]-1,At[i][0]-1])-Wr_tF[At[i][1]-1,At[i][1]-1]])<=sigma[i,0]]
    prob=cvx.Problem(obj,con)
    mydict = {}
    #mydict = {"MSK_DPAR_BASIS_REL_TOL_S":1e-0,"MSK_DPAR_BASIS_TOL_S":1e-0,"MSK_DPAR_BASIS_TOL_X":1e-0,"MSK_IPAR_BI_IGNORE_NUM_ERROR":1}  
    prob.solve(solver='MOSEK', mosek_params=mydict)
    
    M=W.value
    T_tF=u*T_tF
    tF+=1
    print('-----第'+str(tF)+'次迭代罚项大小------')
    print(T_tF * np.sum(sigma.value))
    print('-----第'+str(tF)+'次迭代总生产成本大小------')
    print(p.value)
    Usimilar,SingularValue=getUsimilar(M)
    print('-----第'+str(tF)+'次奇异值比值大小------')
    print(SingularValue[0]/SingularValue[1])
    #print('-----第'+str(tF)+'次权重大小------')
    #print(T_tF)
    #if SingularValue[0]/SingularValue[1]>1e9 and T_tF*np.sum(sigma.value)<1e-1:
        #print("精度达到要求，输出并进行潮流计算")
        #U,Angle=getU_Angle(Usimilar)
        #Pinjection=node_gene@Pg.value-LP
        #print("Pinjection:",Pinjection)
        #Qinjection=node_gene@Qg.value-LQ
        #print("Qinjection:",Qinjection)
        #OutData(U,Angle,Pinjection,Qinjection,menu3,1)
        #PowerFlowCal(Flag_Out)
        #print("工作完成")
        #break
    if  math.log(SingularValue[0]/SingularValue[1],10)>=3:
        print("精度达到要求，输出并进行潮流计算")
        U,Angle=getU_Angle(Usimilar)
        Pinjection=node_gene@Pg.value-LP
        Qinjection=node_gene@Qg.value-LQ
        OutData(U,Angle,Pinjection,Qinjection,menu3,1)
        PowerFlowCal(Flag_Out)
        print("工作完成")
        break
    #print('------pij------')
    #print(Pij.value * Sb)
    #print('------qij------')
    #print(Qij.value * Sb)

    #print('------fai_ij------')
    #print(fai_ij.value)

    #print('-----pij+pji------')
    #print((Pij.value + Pji.value) * Sb)


    # 不等号左边为矩阵形式


